import re
import math
import openpyxl as xl

# parameters
days = []
is_24_hour_format = False
start = 0
end = 0
activity_length = 180
block_length = 5
num_choices = 5

num_days = 0
num_blocks = 0
num_check_blocks = 0
num_members = 0
members = []
scheds = []
choices = []

wb = None # .xlsx workbook

# converts time into array indices for processing
def convert_num_to_time(num):
	if is_24_hour_format:
		hh, mm = num//60, num % 60
		return '{}:{:02}'.format(hh, mm)
	else:
		hh, mm = num//60, num % 60
		meridian = 'AM'
		if hh >= 12:
			meridian = 'PM'
			hh -= 12
		if hh == 0:
			hh = 12
		return '{}:{:02} {}'.format(hh, mm, meridian)

# converts array indices back to time for outputting
def convert_time_to_num(tme):
	if is_24_hour_format:
		hh, mm = map(int, re.split('\W', tme))
		return 60*hh + mm
	else:
		hh, mm, meridian = re.split('\W', tme)
		hh, mm = int(hh), int(mm)
		hh %= 12
		if meridian == 'PM':
			hh += 12
		return 60*hh + mm

# read input from .xlsx file
def read_input():
	global wb
	settings = wb['Settings']
	sheet_scheds = wb['Schedules']

	# reading parameters
	# which days of the week are being considered for scheduling an activity?
	global num_days, days
	num_days = settings.max_column - 1
	last_col = chr(ord('A') + num_days)
	days = list(map(lambda cell: cell.value, settings['B1' : last_col+'1'][0]))

	# 12- or 24-hour format?
	global is_24_hour_format
	yes_or_no = settings['B2'].value
	if yes_or_no.upper() == 'Y':
		is_24_hour_format = True
	else:
		is_24_hour_format = False

	# what times can you schedule the activity?
	global start, end
	ends = re.split('-|–|—', settings['B3'].value)
	ends = list(map(str.strip, ends))
	start = convert_time_to_num(ends[0])
	end   = convert_time_to_num(ends[1])

	# how long is the activity?
	# how efficient can you make the program run?
	# how many best choices do you want to be displayed?
	global activity_length, block_length, num_choices
	activity_length = int(settings['B4'].value)
	block_length = int(settings['B5'].value)
	num_choices = int(settings['B6'].value)

	# processing how many time intervals to make in the array
	global num_blocks, num_check_blocks
	num_blocks = (end - start) // block_length
	num_check_blocks = activity_length // block_length

	# creating arrays of which members are unavailable for each time block for each day
	global num_members, members, scheds
	num_members = math.ceil(sheet_scheds.max_row/(num_days+1))
	members = []	# names
	scheds = []		# 3D arr, 2D arr = day of week
					# 1D_arr[block_idx] = [members unavailable]
	for i in range(num_days):
		scheds.append([[] for i in range(num_blocks)])

	for i in range(num_members):
		members.append(sheet_scheds.cell(row = (num_days+1)*i + 1, column = 1).value)
		for day in range(num_days):
			chosen_row = i*(num_days+1) + day + 2
			cells = sheet_scheds.iter_cols(min_row = chosen_row, max_row = chosen_row,
					min_col = 2, max_col = sheet_scheds.max_column)
			for interval in cells:
				if interval[0].value == None: break
				L, R = map(str.strip, re.split('-|–|—', interval[0].value))
				L, R = map(convert_time_to_num, (L, R))
				L = max((L-start)//block_length, 0)
				R = min((R-start)//block_length, num_blocks)
				for block in range(L, R):
					scheds[day][block].append(i)

# processing input and printing output
def find_available():

	# elements to be sorted by least number of people in conflict with schedule
	global choices
	choices = [] # (num people unavailable, day of week, start time, [people unavailable])

	global num_days, num_blocks, num_check_blocks, scheds
	for day in range(num_days):
		for start_block in range(num_blocks-num_check_blocks+1): # iterate over all possible potential schedules
			unavails = set()
			for block in range(start_block, start_block+num_check_blocks):
				unavails.update(scheds[day][block])
			num_unavail = len(unavails)
			cand_start = start_block * block_length + start
			choices.append((num_unavail, day, cand_start, list(unavails)))

	choices.sort() # find least-conflicting schedules

	# printing output to new worksheet
	global wb
	try:
		del wb['Output']
	except:
		pass
	wb.create_sheet('Output') # clean slate
	output = wb['Output']

	# for making sure input is parsed correctly in 'Schedules' worksheet
	global num_members, days
	output.append(('Number of days detected:', num_days))
	output.append(('Number of members detected:', num_members))
	output.append(())

	# output the best choices
	for choice_num in range(num_choices):
		num_unavail, day, choice_start, unavails = choices[choice_num]
		choice_end = choice_start + activity_length
		output.append((f'Choice #{choice_num+1}:', f'{days[day]}',
			f'{convert_num_to_time(choice_start)} - {convert_num_to_time(choice_end)}'))
		output.append(('Members available:', f'{num_members-num_unavail}/{num_members}'))

		output_list = ['Members unavailable:']
		for i in range(len(unavails)):
			if len(output_list) == 6:
				output.append(output_list)
				output_list = [None]
			output_list.append(members[unavails[i]])
		if len(output_list) > 1:
			output.append(output_list)
		output.append(())
		output.append(())

if __name__ == '__main__':
	wb = xl.load_workbook('schedules.xlsx')
	read_input()
	find_available()

	# formatting output to fit characters
	for col in wb['Output'].columns:
		max_length = 0
		column = chr(ord('A') + col[0].column - 1)
		for cell in col:
			try:
				cell_length = len(str(cell.value))
				max_length = max(max_length, cell_length)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		wb['Output'].column_dimensions[column].width = adjusted_width

	wb.save('schedules.xlsx')
